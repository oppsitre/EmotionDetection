import json
import jieba
import os
import matplotlib.pyplot as plt
import math
from sklearn.cluster import KMeans
import numpy as np
def is_chinese(str):
	for uchar in str:
		if not ('\u4e00' <= uchar<='\u9fff'):
			return False
	return True

# 判断一个unicode是否是数字
def is_number(uchar):
    if '\u0030' <= uchar <='\u0039':
        return True
    else:
        return False

# 判断一个unicode是否是英文字母
def is_alphabet(uchar):
    if ('\u0041' <= uchar<='\u005a') or ('\u0061' <= uchar<='\u007a'):
        return True
    else:
        return False

# 判断是否非汉字，数字和英文字符
def is_other(uchar):
    if not (is_chinese(uchar) or is_number(uchar) or is_alphabet(uchar)):
        return True
    else:
        return False


class Word:
    def __init__(self, content, style):
        self.content = content
        self.style = style

def expressions_read():
    exp_list = []
    with open('../dictionary/expression/expression_labeled_1.txt') as f:
        for line in f:
            line = line.strip().split('\t')
            exp_list.append(line[0])
    return exp_list

def stopwords_read():
	stopwords = []
	with open('../dictionary/word/stopwords.txt') as f:
		for line in f:
			line = line.strip()
			stopwords.append(line)
	return stopwords

def sentence_process(sentence, stopwords):
	sentence_new = []
	for s in sentence:
		if type(s) is tuple:
			sentence_new.append(s[0])
			continue
		seg_list = jieba.cut(s, cut_all=False)
		for seg in seg_list:
			if seg in stopwords:
				continue
			if not is_chinese(seg):
				continue
			sentence_new.append(seg)
	return sentence_new

from openpyxl import load_workbook
def emotion_dict_read():
	wb = load_workbook(filename='../dictionary/word/emotion_dic.xlsx') #载入工作簿
	name2col = { '词语': 'A',
			'词性种类': 'B',
			'词义数': 'C',
			'词义序号': 'D',
			'情感分类': 'E',
			'强度': 'F',
			'极性': 'G',
			'辅助情感分类': 'H',
			'强度': 'I',
			'辅助极性': 'J'
			}
	emtion2id = {'PA': 0, 'PE': 1, 'PD': 2, 'PH': 3, 'PG': 4, 'PB': 5, 'PK': 6,
	 			 'NA': 7, 'NB': 8, 'NJ': 9, 'NH':10, 'PF':11, 'NI':12, 'NC':13,
				 'NG':14, 'NE':15, 'ND':16, 'NN':17, 'NK':18, 'NL':19, 'PC':20
	}

	ws = wb['Sheet1']  #选中工作表
	rows = ws.rows
	columns = ws.columns

	# 行迭代
	content = {}
	for i, row in enumerate(rows):
		if i == 0:
			continue
		line = [col.value for col in row]
		content[line[0]] = [emtion2id[line[4].strip()]]
	return content
	# print(content)
	# print(ws.cell(0,1).value)
def expression_segment(str, expression):
    '''
    The function aims to split word and expressions
    str: the string you want to split
    expression: the list contain all expressions in the dictionary
    Return: a string list and a expression list.
    For example,
    Input: str = 我很开心(^_^)出去玩^_^, expression = [^_^,(^_^),...]
    Output: str_list = ['我很开心', '出去玩'], exp_list = [(^_^), ^_^]
    '''
    pos = 0
    str_tmp = ''
    sentence = []
    while pos < len(str):
        max_len = 0
        exp_tmp = None
        for exp in expression:
            if int(len(str)) - pos < int(len(exp)):
                continue
            if len(exp) <= max_len:
                continue
            if str[pos:(pos+len(exp))] == exp:
                max_len = len(exp)
                exp_tmp = exp
        if max_len > 0:
            if len(str_tmp) > 0:
                sentence.append(str_tmp)
            pos += max_len
            str_tmp = ''
            sentence.append((exp_tmp, 1))

        else:
            str_tmp = str_tmp + str[pos]
            pos += 1

    if len(str_tmp) > 0:
        sentence.append(str_tmp)
    # print('expression_segment')
    return sentence


class Comment:
    def __init__(self, aid, dataset_dir):
        '''
        aid: the av id of each video
        dataset_dir:  the directory of dataset.
        '''
        self.aid = aid
        self.dataset_dir = dataset_dir
        if not self.dataset_dir.endswith('/'):
            self.dataset_dir += '/'
        self.content = self.read(aid)
        '''
        For example, self.content = [[Comment], [Comment], ...]
        '''
    def dfs(self, root):
        cmts = []
        if root is None:
            return None
        for son in root:
            cmts.append(son['content']['message'])
            tmp = self.dfs(son['replies'])
            if tmp != None: cmts.extend(tmp)
        return cmts

    def read(self, aid):
        with open(self.dataset_dir + aid + '/' + aid + '.cmt', 'r') as f:
            cmt = json.load(f)
        cmts = self.dfs(cmt['data']['replies'])
        return cmts

class Danmuku:
	def __init__(self, aid, dataset_dir):
		'''
		aid: the av id of each video
		dataset_dir:  the directory of dataset.
		'''
		self.aid = aid
		self.dataset_dir = dataset_dir
		if not self.dataset_dir.endswith('/'):
		    self.dataset_dir += '/'
		self.content = self.read(aid)
		'''
		For example, self.content = [['danmu', offset in the video, the absolute time published],...]
		'''

	def read(self, aid):
		dan = []
		with open(self.dataset_dir + aid + '/' + aid + '.dan', 'r') as f:
			for line in f:
				line = line.strip().split('\t')
				dan.append((line[0], float(line[1]), line[2]))
		return dan

	def sort(self, content):
		con = sorted(content, key = lambda x: x[1])
		return con

	def split(self, content = None, part_num = 50):
		if content == None:
			content = self.content
		content = self.sort(content)
		min_t = content[0][1]
		max_t = content[-1][1]
		step = (max_t - min_t) / part_num
		pos = min_t
		parts = [0 for i in range(part_num)]
		tmp = 0
		idx = 0
		for c in content:
			if c[1] < pos + step:
				tmp += 1
			else:
				parts[idx] = tmp
				idx += 1
				tmp = 0
				pos = pos + step
		print(len(parts), parts)
		plt.plot(range(part_num), parts)
		plt.show()

	def cluster(self, content = None, part_num = 10):
		if content == None:
			content = self.content
		content = self.sort(content)
		offset = np.array([c[1] for c in content])
		offset = np.reshape(offset, (len(offset), 1))
		kmeans = KMeans(n_clusters=part_num).fit(offset)
		order = {}
		idx = 0
		for i in kmeans.labels_:
			if not(i in order):
				order[i] = idx
				idx += 1
		labels = [order[item] for idx, item in enumerate(kmeans.labels_)]
		uni, parts = np.unique(labels, return_counts = True)
		plt.plot(range(part_num), parts)
		plt.show()
def get_all_text():
	num = 0
	with open('texts.txt', 'w') as con:
		list_dirs = os.walk('../dataset/')
		for root, dirs, files in list_dirs:
			for f in files:
				print(os.path.join(root, f))
				if f.endswith('.dan'):
					dan = Danmuku(f[:-4], '../dataset/')
					# print(dan.content)
					for c in dan.content:
						# print(c)
						con.write(c[0] + '\n')
					num += 1
				if f.endswith('.cmt'):
					dan = Comment(f[:-4], '../dataset/')
					for c in dan.content:
						con.write(c + '\n')
					num += 1

	print('Num:', num)

if __name__ == '__main__':
    # cmt = Comment('4704057', '../dataset/')
    # print(cmt.content)
    # dan = Danmuku('4704057', '../dataset/')
    # str = '我很开心^_^出去玩^_^'
    # exp_list = express_read()
    # sentence = expression_extract(str, exp_list)
    # print(sentence)
	# v = Word('我是', 1)
	# print(type(v) is Word)
	# emotion_dic_read()
	# get_all_text()
	process()
